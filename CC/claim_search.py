"""Utilities for searching similar claims in Excel files."""

from __future__ import annotations

from typing import Any, Dict, List
from difflib import SequenceMatcher

from openpyxl import load_workbook


class ClaimSearcher:
    """Search an Excel workbook for complaints similar to a given text."""

    def __init__(self, excel_path: str, sheet: str | None = None) -> None:
        """Load claim data from ``excel_path``.

        Parameters
        ----------
        excel_path : str
            Path to the Excel file containing past complaints.
        sheet : str, optional
            Name of the sheet to read. Defaults to the active sheet.
        """
        wb = load_workbook(excel_path, read_only=True, data_only=True)
        ws = wb[sheet] if sheet else wb.active
        rows = list(ws.iter_rows(values_only=True))
        wb.close()

        if rows:
            self.headers = [str(v) if v is not None else "" for v in rows[0]]
            self.rows = [tuple(row) for row in rows[1:]]
        else:
            self.headers = []
            self.rows = []

    def _similarity(self, a: str, b: str) -> float:
        """Return a similarity ratio for two strings."""
        return SequenceMatcher(None, a.lower(), b.lower()).ratio()

    def find_similar(self, complaint: str, threshold: float = 0.6) -> List[Dict[str, Any]]:
        """Return rows with complaint text similar to ``complaint``.

        Parameters
        ----------
        complaint : str
            New complaint text to compare.
        threshold : float, optional
            Minimum similarity ratio required for a match. Defaults to ``0.6``.

        Returns
        -------
        List[Dict[str, Any]]
            Matching rows as dictionaries mapping headers to values.
        """
        matches: List[Dict[str, Any]] = []
        for row in self.rows:
            first_cell = str(row[0]) if row and row[0] is not None else ""
            if self._similarity(complaint, first_cell) >= threshold:
                matches.append({h: row[i] for i, h in enumerate(self.headers)})
        return matches
